from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any
import re

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = PROJECT_ROOT / "data"
REPORT_PATH = PROJECT_ROOT / "docs" / "EXCEL_STRUCTURE_REPORT.md"

REQUESTED_FILES = [
    "BoTaiLieu_NhaTre_12-36T_CT388_2026-2027_DieuChinh.xlsx",
    "BoTaiLieu_MauGiao_3-6T_CT388_2026-2027.xlsx",
]

FALLBACK_KEYWORDS = {
    "BoTaiLieu_NhaTre_12-36T_CT388_2026-2027_DieuChinh.xlsx": ["nhatre", "nha tre", "12-36"],
    "BoTaiLieu_MauGiao_3-6T_CT388_2026-2027.xlsx": ["maugiao", "mau giao", "3-6"],
}


@dataclass
class WorkbookTarget:
    requested_name: str
    path: Path | None
    note: str


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\n", " ").replace("\r", " ").strip()
    return re.sub(r"\s+", " ", text)


def normalize_filename(name: str) -> str:
    lowered = name.lower()
    return re.sub(r"[^a-z0-9-]+", "", lowered)


def escape_markdown(value: Any) -> str:
    text = normalize_text(value)
    text = text.replace("|", "\\|")
    return text if text else " "


def trim_cell(value: Any, max_len: int = 120) -> str:
    text = escape_markdown(value)
    if len(text) <= max_len:
        return text
    return text[: max_len - 3].rstrip() + "..."


def row_has_content(row: tuple[Any, ...]) -> bool:
    return any(normalize_text(cell) for cell in row)


def detect_header_row(rows: list[tuple[Any, ...]]) -> int | None:
    best_index: int | None = None
    best_score = -1
    for index, row in enumerate(rows):
        non_empty = [normalize_text(cell) for cell in row if normalize_text(cell)]
        if not non_empty:
            continue
        distinct = len(set(non_empty))
        keyword_score = sum(
            1
            for cell in non_empty
            if any(
                keyword in cell.lower()
                for keyword in [
                    "mã",
                    "ma",
                    "code",
                    "tên",
                    "ten",
                    "nội dung",
                    "noi dung",
                    "mục tiêu",
                    "muc tieu",
                    "yêu cầu",
                    "yeu cau",
                    "lĩnh vực",
                    "linh vuc",
                    "độ tuổi",
                    "do tuoi",
                ]
            )
        )
        score = (len(non_empty) * 2) + distinct + keyword_score
        if score > best_score:
            best_score = score
            best_index = index
    return best_index


def dedupe_columns(values: list[Any]) -> list[str]:
    columns: list[str] = []
    seen: dict[str, int] = {}
    for index, value in enumerate(values, start=1):
        name = normalize_text(value) or f"Column {index}"
        if name in seen:
            seen[name] += 1
            name = f"{name} ({seen[name]})"
        else:
            seen[name] = 1
        columns.append(name)
    return columns


def markdown_table(rows: list[list[Any]], max_cols: int = 12) -> str:
    if not rows:
        return "_Không có dữ liệu trong 10 dòng đầu._"

    col_count = min(max(len(row) for row in rows), max_cols)
    headers = [f"C{i}" for i in range(1, col_count + 1)]
    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join(["---"] * col_count) + " |",
    ]
    for row in rows:
        padded = list(row[:col_count]) + [""] * max(0, col_count - len(row))
        lines.append("| " + " | ".join(trim_cell(cell) for cell in padded) + " |")
    return "\n".join(lines)


def infer_usage(sheet_name: str, columns: list[str], preview_rows: list[tuple[Any, ...]]) -> str:
    haystack = " ".join([sheet_name, *columns, *[normalize_text(cell) for row in preview_rows for cell in row]]).lower()
    suggestions: list[str] = []

    checks = [
        (["yccd", "yêu cầu cần đạt", "yeu cau can dat", "mã yc", "ma yc"], "ánh xạ bảng `yccd` và tra cứu mã/YCCĐ"),
        (["lĩnh vực", "linh vuc", "phát triển", "phat trien"], "ánh xạ bảng `domains` hoặc phân loại lĩnh vực phát triển"),
        (["năng lực", "nang luc"], "ánh xạ bảng `competencies`"),
        (["phẩm chất", "pham chat"], "ánh xạ bảng `qualities`"),
        (["mốc", "moc", "biểu hiện", "bieu hien"], "ánh xạ bảng `milestones` hoặc nội dung quan sát tiến bộ"),
        (["hoạt động", "hoat dong", "trò chơi", "tro choi"], "ánh xạ bảng `activities` hoặc ngân hàng hoạt động"),
        (["rubric", "tiêu chí", "tieu chi", "minh chứng", "minh chung"], "ánh xạ bảng `rubrics`, `observations` hoặc `assessments`"),
        (["kế hoạch", "ke hoach", "tháng", "thang", "tuần", "tuan", "ngày", "ngay"], "phục vụ lập `year_plans`, `month_plans`, `week_plans`, `day_plans`"),
        (["nhà trẻ", "nha tre", "12-36"], "dữ liệu nhóm nhà trẻ 12-36 tháng"),
        (["mẫu giáo", "mau giao", "3-6"], "dữ liệu nhóm mẫu giáo 3-6 tuổi"),
    ]
    for keywords, suggestion in checks:
        if any(keyword in haystack for keyword in keywords):
            suggestions.append(suggestion)

    if not suggestions:
        return "Cần rà soát thủ công thêm trước khi ánh xạ; sheet có thể là danh mục/phụ lục hoặc phần trình bày."
    return "; ".join(dict.fromkeys(suggestions)) + "."


def find_targets() -> tuple[list[WorkbookTarget], list[str]]:
    notes: list[str] = []
    targets: list[WorkbookTarget] = []
    available = list(DATA_DIR.glob("*.xlsx"))
    normalized_lookup = {normalize_filename(path.name): path for path in available}
    used_paths: set[Path] = set()

    for requested in REQUESTED_FILES:
        requested_path = DATA_DIR / requested
        if requested_path.exists():
            targets.append(WorkbookTarget(requested, requested_path, "Đúng tên file yêu cầu."))
            used_paths.add(requested_path)
            continue

        missing_note = f"Thiếu file đúng tên: `{requested_path}`."
        notes.append(missing_note)
        requested_norm = normalize_filename(requested)
        fallback_path = normalized_lookup.get(requested_norm)
        if fallback_path is None:
            keywords = FALLBACK_KEYWORDS.get(requested, [])
            for path in available:
                candidate_norm = normalize_filename(path.name)
                if path in used_paths:
                    continue
                if any(normalize_filename(keyword) in candidate_norm for keyword in keywords):
                    fallback_path = path
                    break

        if fallback_path is not None:
            used_paths.add(fallback_path)
            targets.append(
                WorkbookTarget(
                    requested,
                    fallback_path,
                    f"{missing_note} Đã tự dùng file gần đúng hiện có: `{fallback_path.name}`.",
                )
            )
        else:
            targets.append(WorkbookTarget(requested, None, missing_note))

    return targets, notes


def inspect_sheet(workbook_path: Path, sheet_name: str) -> str:
    workbook = load_workbook(workbook_path, read_only=False, data_only=True)
    sheet = workbook[sheet_name]
    row_count = sheet.max_row or 0
    col_count = sheet.max_column or 0
    merged_ranges = [str(merged_range) for merged_range in sheet.merged_cells.ranges]

    preview_15 = [
        tuple(cell.value for cell in row)
        for row in sheet.iter_rows(min_row=1, max_row=min(15, row_count), max_col=col_count)
    ]
    preview_10_rows = [
        list(row)
        for row in preview_15[:10]
        if row_has_content(row)
    ]

    header_index = detect_header_row(preview_15)
    columns = dedupe_columns(list(preview_15[header_index])) if header_index is not None else []
    usage = infer_usage(sheet_name, columns, preview_15)

    lines = [
        f"### Sheet: `{sheet_name}`",
        "",
        f"- Tên file: `{workbook_path.name}`",
        f"- Số dòng: `{row_count}`",
        f"- Số cột: `{col_count}`",
        f"- Dòng tiêu đề dự đoán: `{header_index + 1 if header_index is not None else 'chưa xác định'}`",
        f"- Có merged cells: `{'Có' if merged_ranges else 'Không'}`",
    ]
    if merged_ranges:
        lines.append(f"- Merged ranges phát hiện: `{', '.join(merged_ranges[:20])}`")
        if len(merged_ranges) > 20:
            lines.append(f"- Merged ranges còn lại: `{len(merged_ranges) - 20}`")
    lines.extend(
        [
            "",
            "**Danh sách cột phát hiện được**",
            "",
            ", ".join(f"`{column}`" for column in columns) if columns else "_Chưa xác định được cột từ 15 dòng đầu._",
            "",
            "**10 dòng đầu tiên**",
            "",
            markdown_table(preview_10_rows),
            "",
            "**Nhận xét sơ bộ**",
            "",
            usage,
            "",
        ]
    )
    workbook.close()
    return "\n".join(lines)


def build_report() -> str:
    targets, missing_notes = find_targets()
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    lines = [
        "# EXCEL_STRUCTURE_REPORT.md",
        "",
        "## Báo cáo phân tích cấu trúc Excel CT388",
        "",
        f"- Thời điểm tạo: `{generated_at}`",
        f"- Thư mục dữ liệu: `{DATA_DIR}`",
        "- Phạm vi: chỉ đọc cấu trúc workbook và 15 dòng đầu mỗi sheet; không import vào database.",
        "",
    ]

    if missing_notes:
        lines.extend(["## Cảnh báo file", ""])
        for note in missing_notes:
            lines.append(f"- {note}")
        lines.append("")

    for target in targets:
        lines.extend(
            [
                f"## Workbook yêu cầu: `{target.requested_name}`",
                "",
                f"- Ghi chú: {target.note}",
                "",
            ]
        )
        if target.path is None:
            lines.append("_Không thể phân tích vì chưa tìm thấy file thay thế phù hợp._")
            lines.append("")
            continue

        workbook = load_workbook(target.path, read_only=False, data_only=True)
        sheet_names = workbook.sheetnames
        workbook.close()
        lines.append(f"- File phân tích thực tế: `{target.path}`")
        lines.append(f"- Số sheet: `{len(sheet_names)}`")
        lines.append("")
        for sheet_name in sheet_names:
            lines.append(inspect_sheet(target.path, sheet_name))

    lines.extend(
        [
            "## Ghi chú kỹ thuật",
            "",
            "- Báo cáo này chưa thực hiện import Excel vào SQLite.",
            "- Báo cáo này không sửa mã/YCCĐ hoặc nội dung giáo dục.",
            "- Các nhận xét chức năng là gợi ý sơ bộ dựa trên tên sheet, tên cột và dữ liệu xem trước; cần xác nhận trước khi thiết kế mapping import ở bước sau.",
            "",
        ]
    )
    return "\n".join(lines)


def main() -> None:
    REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
    report = build_report()
    REPORT_PATH.write_text(report, encoding="utf-8")
    print(f"Excel structure report written to: {REPORT_PATH}")


if __name__ == "__main__":
    main()
