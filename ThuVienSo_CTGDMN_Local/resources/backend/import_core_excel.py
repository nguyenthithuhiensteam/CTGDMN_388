from __future__ import annotations
from collections import Counter,defaultdict
from datetime import datetime
from pathlib import Path
import re
import sys

ROOT=Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
 sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook
from backend import db
from backend.init_db import TABLES,init_database
try:
 sys.stdout.reconfigure(encoding='utf-8', errors='replace')
except Exception:
 pass
DATA=ROOT/'data'; REPORT=ROOT/'docs'/'IMPORT_CORE_REPORT.md'
MG=DATA/'BoTaiLieu_MauGiao_3-6T_CT388_2026-2027.xlsx'; NT=DATA/'BoTaiLieu_NhaTre_12-36T_CT388_2026-2027_DieuChinh.xlsx'; YEAR='2026-2027'
CORE=['rubrics','activities','milestones','year_plans','yccd','competencies','qualities','domains','age_groups']
# "8. NGÂN HÀNG HOẠT ĐỘNG" thực ra chứa 6 bảng khác nhau ghép dọc trong cùng 1
# sheet (bảng chính NT-xx ở đầu + 5 bảng "HOẠT ĐỘNG BỔ SUNG THEO MÃ" — mỗi bảng
# 1 lĩnh vực TC/TX/NN/NT/NgT — có tiêu đề cột KHÁC bảng chính). Trước đây cả sheet
# bị đọc bằng 1 header cố định (dòng 5) nên 5 bảng bổ sung bị đọc sai cột — cột
# "Thời lượng gợi ý" (vd "10 phút") của chúng bị hiểu nhầm thành "Lĩnh vực chính",
# tạo ra các dòng domains rác. Khai báo riêng từng khối với đúng header + biên
# dòng (end_row) để đọc đúng, không chạm vào bảng chính (dòng 6-16).
NT_HD_EXTRA_DOMAIN={'nt_hd_extra_tc':'TC','nt_hd_extra_tx':'TX','nt_hd_extra_nn':'NN','nt_hd_extra_nt':'NT','nt_hd_extra_ngt':'NgT'}
CONFIGS=[(MG,'KHUNGNL',4,'mg_khung'),(MG,'MATRANYCCD',4,'mg_yccd'),(MG,'NGANHANGHD',4,'mg_hd'),(MG,'RUBRIC',5,'mg_rubric'),(MG,'KHNAM_3_4T',8,'mg_year'),(MG,'KHNAM_4_5T',8,'mg_year'),(MG,'KHNAM_5_6T',8,'mg_year'),
 (NT,'6. CẦU NỐI KHUNG NL',5,'nt_bridge'),
 (NT,'8. NGÂN HÀNG HOẠT ĐỘNG',5,'nt_hd',16),
 (NT,'8. NGÂN HÀNG HOẠT ĐỘNG',18,'nt_hd_extra_tc',38),
 (NT,'8. NGÂN HÀNG HOẠT ĐỘNG',40,'nt_hd_extra_tx',63),
 (NT,'8. NGÂN HÀNG HOẠT ĐỘNG',65,'nt_hd_extra_nn',80),
 (NT,'8. NGÂN HÀNG HOẠT ĐỘNG',82,'nt_hd_extra_nt',99),
 (NT,'8. NGÂN HÀNG HOẠT ĐỘNG',101,'nt_hd_extra_ngt',112),
 (NT,'9b. GIÁO DỤC TRONG SINH HOẠT',5,'nt_daily'),(NT,'11. RUBRIC NHÀ TRẺ',6,'nt_rubric'),
 # Chạy SAU '11. RUBRIC NHÀ TRẺ': yccd theo mã đã được tạo với age_group_id=
 # '24-36th' ở bước đó; FRAMEWORK không truyền lại age_group_id (None) nên nhờ
 # coalesce trong upsert_returning_id mà giá trị đó được GIỮ NGUYÊN — FRAMEWORK
 # chỉ bổ sung milestones theo 8 mốc (dữ liệu 12-36 tháng còn thiếu hoàn toàn
 # trước đây) và làm giàu source_note của yccd, không ghi đè liên kết cũ.
 (NT,'2. FRAMEWORK',4,'nt_framework')]
REQ={'mg_khung':['Lĩnh vực','Mã','Tên năng lực'],'mg_yccd':['Mã','YCCĐ 388 (mức M - đích chung 5-6 tuổi)'],'mg_hd':['Mã HĐ','Độ tuổi','Chủ đề/Chủ đề nhánh','Lĩnh vực chính','Mục tiêu quan sát được'],'mg_rubric':['Lĩnh vực','Mã','Tên năng lực','Biểu hiện I','Biểu hiện H','Biểu hiện V'],'mg_year':['Lĩnh vực','Mã','Mục tiêu năm học'],'nt_bridge':['Lĩnh vực Nhà trẻ (82 chỉ số)','Năng lực nền tảng tương ứng','Phẩm chất tương ứng'],'nt_framework':['Mã','Năng lực','YCCĐ 36th'],'nt_hd':['Mã','Hoạt động gợi ý (qua chơi)','Lĩnh vực chính'],'nt_daily':['Thời điểm sinh hoạt','Giáo dục lồng ghép (gợi ý)','Mã năng lực liên quan'],'nt_rubric':['Lĩnh vực','Mã năng lực','I – Khởi đầu','H – Hình thành','V – Vững chắc (đạt đích 36 tháng)']}
for _k in NT_HD_EXTRA_DOMAIN: REQ[_k]=['Mã','Hoạt động ngắn – nhẹ','Bối cảnh','Thời lượng gợi ý']
MANUAL=['Nhà trẻ: 1. ĐỒNG BỘ KẾ HOẠCH','Nhà trẻ: 3. THƯ VIỆN SỐ','Nhà trẻ: 4. TRA CỨU NHANH','Nhà trẻ: 7. KHUNG NL TÓM TẮT','Mẫu giáo: MOCPT','Mẫu giáo: PHIENCHE','Hồ sơ cá nhân/portfolio/bảng kiểm lớp/nhật ký quan sát']
GAPS=['Thiếu bảng nối activity-competency, activity-quality, plan-yccd.','milestones chưa có competency_id/yccd_id nên mốc đang lưu gián tiếp qua title/description.','activities thiếu cột riêng cho câu hỏi, biểu hiện, minh chứng, hòa nhập, mở rộng; giữ trong notes.','year_plans thiếu bảng chi tiết mục tiêu năm theo mã; mỗi mục tiêu lưu thành một dòng.']
def t(v): return re.sub(r'\s+',' ',str(v).replace('\n',' ').replace('\r',' ').strip()) if v is not None else ''
def code(v): return t(v).replace('–','-').replace('—','-')
def split(v):
 s=t(v)
 return [] if not s or s in {'—','-'} else [x.strip() for x in re.split(r'[,;·]+',s) if x.strip() and x.strip() not in {'—','-'}]
def age(v):
 s=t(v).lower()
 return 'MGB' if '3-4' in s or 'mgb' in s else 'MGN' if '4-5' in s or 'mgn' in s else 'MGL' if '5-6' in s or 'mgl' in s else '12-24th' if '12' in s and '24' in s else '24-36th' if '24' in s and '36' in s else None
def dparse(v):
 s=t(v); m=re.match(r'^([A-Za-zÀ-ỹ]+)\s*[–\-]\s*(.+)$',s)
 return (m.group(1),m.group(2)) if m and len(m.group(1))<=4 else (None,s)
def cparse(v):
 s=t(v); m=re.match(r'^([A-Za-zÀ-ỹ]+\d+(?:\.\d+)?)\s*[–\-]\s*(.+)$',s)
 if m: return code(m.group(1)),m.group(2)
 m=re.match(r'^([A-Za-zÀ-ỹ]+\d+(?:\.\d+)?)',s); return (code(m.group(1)),s) if m else ('',s)
def gid(conn,table,col,val):
 if not val: return None
 row=conn.execute(table.select().filter_by(**{col:val})).mappings().fetchone(); return int(row['id']) if row else None
def find_or_upsert(conn,table,match,values):
 """Idempotent insert-or-update for tables with no natural unique `code` column
 (milestones/rubrics/year_plans): look up an existing row by `match` (a dict of
 columns that together identify "the same logical row" across re-imports), and
 UPDATE it in place if found (keeping its id), else INSERT a new row. This is what
 lets the importer run repeatedly without a destructive DELETE-then-reinsert."""
 clean={k:v for k,v in match.items() if v is not None}
 existing=conn.execute(table.select().filter_by(**clean)).mappings().fetchone() if clean else None
 if existing:
  conn.execute(table.update().where(table.c.id==existing['id']).values(**values))
  return int(existing['id'])
 return db.insert_returning_id(conn,table,{**match,**values})
def domain_by_code(conn,dc):
 dc=code(dc)
 return gid(conn,db.domains,'code',dc) if dc else None
def domain_smart(conn,raw):
 """Resolve 1 lĩnh vực từ text tự do: 'TC – Thể chất' (có mã), mã trần 'TX' (nt_hd
 hay ghi tắt kiểu này), hoặc tên đầy đủ. Ưu tiên mã đã có (tránh tạo domain trùng
 tên/không mã như 'TX'/'TC' từng bị tạo khi chỉ so khớp theo tên)."""
 s=t(raw)
 if not s: return None
 dc,dn=dparse(s)
 if dc: return domain(conn,dn,dc)
 existing=domain_by_code(conn,s)
 return existing if existing else domain(conn,s)
def domain(conn,name,dc=None,desc=''):
 name=t(name); dc=t(dc)
 if not name: return None
 if dc:
  return db.upsert_returning_id(conn,db.domains,{'code':dc,'name':name,'description':desc},['code'])
 old=gid(conn,db.domains,'name',name)
 return old or db.insert_returning_id(conn,db.domains,{'name':name,'description':desc})
def comp(conn,cc,name,desc='',did=None):
 cc=code(cc); name=t(name) or cc
 if not name and not cc: return None
 if cc:
  return db.upsert_returning_id(conn,db.competencies,{'code':cc,'name':name,'description':desc,'domain_id':did},['code'],coalesce_cols=['domain_id'])
 old=gid(conn,db.competencies,'name',name)
 return old or db.insert_returning_id(conn,db.competencies,{'name':name,'description':desc,'domain_id':did})
def qual(conn,name,desc=''):
 name=t(name)
 if not name: return None
 old=gid(conn,db.qualities,'name',name)
 return old or db.insert_returning_id(conn,db.qualities,{'name':name,'description':desc})
def yccd(conn,cc,content,aid=None,did=None,cid=None,note=''):
 cc=code(cc); content=t(content)
 if not cc or not content: return None
 return db.upsert_returning_id(conn,db.yccd,{'code':cc,'content':content,'age_group_id':aid,'domain_id':did,'competency_id':cid,'source_note':note},['code'],coalesce_cols=['age_group_id','domain_id','competency_id'])
def milestone(conn,title,desc,aid=None,did=None,evi=''):
 title=t(title); desc=t(desc)
 if not title or not desc: return None
 return find_or_upsert(conn,db.milestones,{'title':title},{'age_group_id':aid,'domain_id':did,'description':desc,'evidence_hint':evi,'updated_at':db.now_str()})
def activity(conn,cc,title,aid=None,did=None,yid=None,obj='',mat='',steps='',notes=''):
 cc=t(cc); title=t(title)
 if not title: return None
 if cc:
  return db.upsert_returning_id(conn,db.activities,{'code':cc,'title':title,'age_group_id':aid,'domain_id':did,'yccd_id':yid,'objective':obj,'materials':mat,'steps':steps,'notes':notes},['code'],coalesce_cols=['age_group_id','domain_id','yccd_id'])
 return find_or_upsert(conn,db.activities,{'title':title},{'age_group_id':aid,'domain_id':did,'yccd_id':yid,'objective':obj,'materials':mat,'steps':steps,'notes':notes,'updated_at':db.now_str()})
def rubric(conn,title,crit,yid=None,evi='',sup=''):
 title=t(title)
 if not title or not t(crit): return None
 return find_or_upsert(conn,db.rubrics,{'title':title},{'yccd_id':yid,'criteria':crit,'evidence_hint':evi,'support_next':sup,'updated_at':db.now_str()})
def year(conn,aid,title,notes):
 title=t(title)
 if not title and not t(notes): return None
 return find_or_upsert(conn,db.year_plans,{'title':title,'age_group_id':aid},{'school_year':YEAR,'notes':t(notes),'updated_at':db.now_str()})
def prep(conn):
 # KHÔNG xóa dữ liệu lõi cũ ở đây nữa: mọi bảng lõi (yccd/milestones/activities/
 # rubrics/domains/competencies/qualities/year_plans) đã có FK trỏ tới từ dữ liệu
 # thật của trường (annual_plan_goals.yccd_id, assessments.yccd_id, activities...),
 # nên xóa rồi chèn lại sẽ đổi ID (đặc biệt trên Postgres, nơi sequence không bao
 # giờ lùi lại) và làm gãy các liên kết đó một cách âm thầm. Toàn bộ importer bên
 # dưới chạy qua upsert theo `code` ổn định (hoặc find-or-update theo tiêu đề khi
 # bảng không có cột code) nên chạy lại nhiều lần vẫn an toàn, không tạo trùng.
 for cc,name,desc in [('12-24th','Nhà trẻ 12-24 tháng','Nhóm tuổi nhà trẻ'),('24-36th','Nhà trẻ 24-36 tháng','Nhóm tuổi nhà trẻ'),('MGB','Mẫu giáo bé 3-4 tuổi','3-4 tuổi'),('MGN','Mẫu giáo nhỡ 4-5 tuổi','4-5 tuổi'),('MGL','Mẫu giáo lớn 5-6 tuổi','5-6 tuổi')]:
  db.upsert_returning_id(conn,db.age_groups,{'code':cc,'name':name,'description':desc},['code'])
def row_import(c,k,s,row,skip,rn):
 if k=='mg_khung':
  cc,name,dom=code(row.get('Mã')),row.get('Tên năng lực'),row.get('Lĩnh vực')
  if not(cc and name and dom): return skip(rn,'Thiếu Lĩnh vực/Mã/Tên năng lực')
  did=domain(c,dom); cid=comp(c,cc,name,did=did)
  for q in split(row.get('Phẩm chất')): qual(c,q)
  for col,ac in [('3-4 tuổi (I)','MGB'),('4-5 tuổi (R)','MGN'),('5-6 tuổi (M)','MGL')]:
   if row.get(col): milestone(c,f'{cc} - {col}',row[col],gid(c,db.age_groups,'code',ac),did,row.get('Minh chứng quan sát',''))
  return bool(cid)
 if k=='mg_yccd':
  cc,content=code(row.get('Mã')),row.get('YCCĐ 388 (mức M - đích chung 5-6 tuổi)')
  if not(cc and content): return skip(rn,'Thiếu Mã hoặc YCCĐ 388')
  cid=comp(c,cc,row.get('Tên năng lực') or cc); note=' | '.join(x for x in [row.get('Đối chiếu CS QĐ 4222 (chỉ MGL)'),row.get('Áp dụng độ tuổi'),row.get('Ghi chú')] if x)
  return bool(yccd(c,cc,content,gid(c,db.age_groups,'code','MGL'),cid=cid,note=note))
 if k=='mg_hd':
  cc,title,obj=row.get('Mã HĐ'),row.get('Chủ đề/Chủ đề nhánh'),row.get('Mục tiêu quan sát được')
  if not(cc and title and obj): return skip(rn,'Thiếu Mã HĐ/Chủ đề/Mục tiêu quan sát được')
  did=domain(c,row.get('Lĩnh vực chính')); kc,kn=cparse(row.get('Năng lực lĩnh vực')); comp(c,kc,kn,did=did)
  for q in split(row.get('Phẩm chất liên quan')): qual(c,q)
  notes=' | '.join(x for x in [f"Lĩnh vực tích hợp thêm: {row.get('Lĩnh vực tích hợp thêm','')}",f"Năng lực nền tảng: {row.get('Năng lực nền tảng','')}",f"Câu hỏi gợi mở: {row.get('Câu hỏi gợi mở của GV','')}",f"Biểu hiện mong đợi: {row.get('Biểu hiện mong đợi','')}",f"Minh chứng đánh giá: {row.get('Minh chứng đánh giá','')}",f"Điều chỉnh hòa nhập / hỗ trợ đặc biệt: {row.get('Điều chỉnh hòa nhập / hỗ trợ đặc biệt','')}",f"Mở rộng cho trẻ khá hơn: {row.get('Mở rộng cho trẻ khá hơn','')}"] if not x.endswith(': '))
  return bool(activity(c,cc,title,gid(c,db.age_groups,'code',age(row.get('Độ tuổi')) or ''),did,gid(c,db.yccd,'code',kc),obj,row.get('Chuẩn bị môi trường, học liệu',''),row.get('Tiến trình tổ chức',''),notes))
 if k=='mg_rubric':
  cc,name=code(row.get('Mã')),row.get('Tên năng lực')
  if not(cc and name): return skip(rn,'Thiếu Mã hoặc Tên năng lực')
  did=domain(c,row.get('Lĩnh vực')); comp(c,cc,name,did=did); crit='\n'.join(x for x in [f"I: {row.get('Biểu hiện I','')}",f"H: {row.get('Biểu hiện H','')}",f"V: {row.get('Biểu hiện V','')}"] if not x.endswith(': '))
  return bool(rubric(c,f'{cc} - {name}',crit,gid(c,db.yccd,'code',cc),row.get('Minh chứng',''),row.get('Hướng hỗ trợ tiếp theo','')))
 if k=='mg_year':
  cc,obj,dom=code(row.get('Mã')),row.get('Mục tiêu năm học'),row.get('Lĩnh vực')
  if not(cc and obj and dom): return skip(rn,'Thiếu Lĩnh vực/Mã/Mục tiêu năm học')
  did=domain(c,dom); comp(c,cc,cc,did=did); ac={'KHNAM_3_4T':'MGB','KHNAM_4_5T':'MGN','KHNAM_5_6T':'MGL'}.get(s)
  return bool(year(c,gid(c,db.age_groups,'code',ac),f'{s} - {dom} - {cc}',obj))
 if k=='nt_bridge':
  raw=row.get('Lĩnh vực Nhà trẻ (82 chỉ số)')
  if not raw: return skip(rn,'Thiếu Lĩnh vực Nhà trẻ')
  if raw.lower().startswith('ghi chú'): return skip(rn,'Dòng ghi chú cuối bảng, không phải lĩnh vực')
  dc,dn=dparse(raw); did=domain(c,dn or raw,dc,row.get('Diễn giải liên thông',''))
  for x in split(row.get('Năng lực nền tảng tương ứng')): comp(c,None,x,row.get('Diễn giải liên thông',''),did)
  for x in split(row.get('Phẩm chất tương ứng')): qual(c,x,row.get('Diễn giải liên thông',''))
  return bool(did)
 if k=='nt_hd':
  cc,title=row.get('Mã'),row.get('Hoạt động gợi ý (qua chơi)')
  if not(cc and title): return skip(rn,'Thiếu Mã hoặc Hoạt động gợi ý')
  first=split(row.get('Lĩnh vực chính'))[:1]; did=domain_smart(c,first[0]) if first else None
  notes=' | '.join(x for x in [f"Chủ đề: {row.get('Chủ đề','')}",f"Lĩnh vực chính: {row.get('Lĩnh vực chính','')}",f"STEAM: {row.get('STEAM','')}",f"SEL: {row.get('SEL','')}"] if not x.endswith(': '))
  return bool(activity(c,cc,title,did=did,notes=notes))
 if k=='nt_daily':
  title,steps=row.get('Thời điểm sinh hoạt'),row.get('Giáo dục lồng ghép (gợi ý)')
  if not(title and steps): return skip(rn,'Thiếu Thời điểm sinh hoạt hoặc Giáo dục lồng ghép')
  notes=' | '.join(x for x in [f"Mã năng lực liên quan: {row.get('Mã năng lực liên quan','')}",f"Lưu ý an toàn/chăm sóc: {row.get('Lưu ý an toàn/chăm sóc','')}"] if not x.endswith(': '))
  return bool(activity(c,None,f'Sinh hoạt hằng ngày - {title}',steps=steps,notes=notes))
 if k=='nt_rubric':
  cc,dom,v=code(row.get('Mã năng lực')),row.get('Lĩnh vực'),row.get('V – Vững chắc (đạt đích 36 tháng)')
  if not(cc and dom and v): return skip(rn,'Thiếu Lĩnh vực/Mã năng lực/Vững chắc')
  did=domain(c,dom); cid=comp(c,cc,cc,did=did); yid=yccd(c,cc,v,gid(c,db.age_groups,'code','24-36th'),did,cid,'Nhà trẻ - RUBRIC NHÀ TRẺ')
  for lv,col in [('I','I – Khởi đầu'),('H','H – Hình thành'),('V','V – Vững chắc (đạt đích 36 tháng)')]:
   if row.get(col): milestone(c,f'{cc} - mức {lv}',row[col],did=did,evi=row.get('Minh chứng quan sát',''))
  crit='\n'.join(x for x in [f"I: {row.get('I – Khởi đầu','')}",f"H: {row.get('H – Hình thành','')}",f'V: {v}'] if not x.endswith(': '))
  return bool(rubric(c,f'{cc} - Rubric nhà trẻ',crit,yid,row.get('Minh chứng quan sát',''),row.get('Hỗ trợ tiếp theo','')))
 if k in NT_HD_EXTRA_DOMAIN:
  cc,title,ctx,dur,sens=row.get('Mã'),row.get('Hoạt động ngắn – nhẹ'),row.get('Bối cảnh'),row.get('Thời lượng gợi ý'),row.get('Đa giác quan')
  if not(cc and title): return skip(rn,'Thiếu Mã hoặc Hoạt động ngắn – nhẹ')
  did=domain_by_code(c,NT_HD_EXTRA_DOMAIN[k])
  notes=' | '.join(x for x in [f'Bối cảnh: {ctx}',f'Thời lượng gợi ý: {dur}',f'Đa giác quan: {sens}'] if not x.endswith(': '))
  return bool(activity(c,cc,title,did=did,notes=notes))
 if k=='nt_framework':
  cc,name,term=code(row.get('Mã')),row.get('Năng lực'),row.get('YCCĐ 36th')
  if not(cc and term): return skip(rn,'Thiếu Mã hoặc YCCĐ 36th')
  m=re.match(r'^([A-Za-zÀ-ỹ]+)',cc); did=domain_by_code(c,m.group(1) if m else None)
  ncc,nname=cparse(name); cid=comp(c,ncc or cc,nname or name,did=did)
  yid=yccd(c,cc,term,did=did,cid=cid,note='Nguồn: sheet "2. FRAMEWORK" (82 mã x 8 mốc) - YCCĐ đích 36 tháng, áp dụng chung cho toàn bộ Nhà trẻ 12-36 tháng (không riêng nhóm 24-36 tháng).')
  evi=row.get('Minh chứng','')
  ok=bool(yid)
  for col,ac in [('NHÓM 12–24 THÁNG 12–15th','12-24th'),('15–18th','12-24th'),('18–21th','12-24th'),('21–24th','12-24th'),('NHÓM 24–36 THÁNG 24–27th','24-36th'),('27–30th','24-36th'),('30–33th','24-36th'),('33–36th','24-36th')]:
   if row.get(col) and milestone(c,f'{cc} - {col}',row[col],gid(c,db.age_groups,'code',ac),did,evi): ok=True
  return ok
 return skip(rn,f'Chưa có importer {k}')
def import_sheet(conn,cfg):
 file,sheet,hr,k=cfg[0],cfg[1],cfg[2],cfg[3]; end_row=cfg[4] if len(cfg)>4 else None
 res={'file':file.name,'sheet':sheet,'rows':0,'ok':0,'skip':0,'warn':[],'reasons':Counter(),'errors':[]}
 def sk(r,why): res['skip']+=1; res['reasons'][why]+=1; res['errors'].append(f'Dòng {r}: {why}'); return False
 if not file.exists(): res['warn'].append(f'Không tìm thấy file: {file}'); return res
 wb=load_workbook(file,data_only=True)
 if sheet not in wb.sheetnames: res['warn'].append(f'Không tìm thấy sheet: {sheet}'); wb.close(); return res
 ws=wb[sheet]; heads=[t(ws.cell(hr,i).value) or f'Column {i}' for i in range(1,ws.max_column+1)]
 miss=[x for x in REQ[k] if x not in heads]
 if miss: res['warn'].append('Thiếu cột bắt buộc: '+', '.join(miss)); wb.close(); return res
 last_row=min(end_row,ws.max_row) if end_row else ws.max_row
 if any(x.startswith('Column ') for x in heads): res['warn'].append('Có cột Column N không rõ nhãn; đã bỏ qua các cột này.')
 for rn in range(hr+1,last_row+1):
  row={h:t(ws.cell(rn,i+1).value) for i,h in enumerate(heads) if h and not h.startswith('Column ')}
  if not any(row.values()): continue
  res['rows']+=1
  try:
   with conn.begin_nested():
    ok=row_import(conn,k,sheet,row,sk,rn)
   if ok: res['ok']+=1
  except Exception as e: sk(rn,f'Lỗi import: {e}')
 wb.close(); return res
def make_report(results,ct,removed=None):
 removed=removed or {}
 lines=['# IMPORT_CORE_REPORT.md','','## Báo cáo import dữ liệu lõi CT388','',f"- Thời điểm import: `{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}`",'- Phạm vi: chỉ import sheet lõi; không import hồ sơ cá nhân, portfolio, bảng kiểm lớp, nhật ký quan sát.','- Cơ chế chạy lại: KHÔNG xóa dữ liệu lõi cũ — mọi bảng ghi qua upsert theo mã ổn định (hoặc find-or-update theo tiêu đề), giữ nguyên id đã có, an toàn khi chạy lại nhiều lần.',f"- Dọn rác từ lỗi import trước đây: xóa `{removed.get('activities',0)}` hoạt động và `{removed.get('domains',0)}` lĩnh vực rác đã xác nhận không còn bị tham chiếu.",'','## Tổng hợp theo sheet','','| File | Sheet | Dòng đọc | Import thành công | Bỏ qua | Cảnh báo |','|---|---|---:|---:|---:|---|']
 for r in results: lines.append(f"| `{r['file']}` | `{r['sheet']}` | {r['rows']} | {r['ok']} | {r['skip']} | {'<br>'.join(r['warn'])} |")
 lines+=['','## Lý do bỏ qua','']
 for r in results:
  if r['reasons']:
   lines.append(f"### `{r['sheet']}`")
   for why,n in r['reasons'].items(): lines.append(f'- {why}: `{n}` dòng')
   for e in r['errors'][:50]: lines.append(f'  - {e}')
 lines+=['','## Danh sách sheet import được','']+[f"- `{r['file']}` / `{r['sheet']}`: `{r['ok']}` dòng" for r in results if r['ok']]
 lines+=['','## Danh sách sheet cần kiểm tra thủ công','']+[f'- {x}' for x in MANUAL]
 lines+=['','## Điểm schema cần cân nhắc trước import sâu','']+[f'- {x}' for x in GAPS]
 lines+=['','## Số bản ghi hiện có trong database','','| Bảng | Số bản ghi |','|---|---:|']
 for k,v in ct.items(): lines.append(f'| `{k}` | {v} |')
 return '\n'.join(lines)+'\n'
# Rác cụ thể do một lỗi đọc sheet "8. NGÂN HÀNG HOẠT ĐỘNG" ở các lần import
# TRƯỚC lần sửa này để lại (sheet có 6 bảng ghép dọc nhưng từng bị đọc bằng 1
# header cố định — xem chú thích ở CONFIGS/NT_HD_EXTRA_DOMAIN phía trên). Chỉ
# xóa đúng các bản ghi khớp CHÍNH XÁC danh sách này (không suy đoán theo id, vì
# id có thể khác nhau giữa các môi trường), và CHỈ xóa domain nếu xác nhận
# KHÔNG còn bảng nào tham chiếu tới — an toàn để chạy lại nhiều lần.
_GARBAGE_ACTIVITY=[('Mã','Bối cảnh')]  # (code, title) của dòng header bị đọc nhầm thành hoạt động
_GARBAGE_DOMAIN_NAMES=['TX','TC','NT','NN','NgT','Thời lượng gợi ý','2-3 phút','3-5 phút','Cả buổi','5 phút','3 phút','1 phút/lần','5-10 phút','Cả bữa ăn','1 phút','Lặp lại hằng ngày','2 phút','Quan sát hằng ngày','10 phút','5-8 phút','Tự do','Ghi chú: Bảng chỉ để liên thông cách gọi tên; đánh giá trẻ nhà trẻ vẫn theo 82 chỉ số & 8 mốc, thang I–H–V.']
def cleanup_known_import_garbage(conn):
 removed={'activities':0,'domains':0}
 for cc,title in _GARBAGE_ACTIVITY:
  row=conn.execute(db.activities.select().filter_by(code=cc,title=title)).mappings().fetchone()
  if row:
   conn.execute(db.activities.delete().where(db.activities.c.id==row['id'])); removed['activities']+=1
 ref_tables=[(db.competencies,'domain_id'),(db.yccd,'domain_id'),(db.activities,'domain_id'),(db.milestones,'domain_id')]
 for name in _GARBAGE_DOMAIN_NAMES:
  # Chỉ nhắm vào bản ghi domain KHÔNG có `code` (bản ghi có code là domain thật,
  # được tạo đúng bằng sheet "6. CẦU NỐI KHUNG NL" — không bao giờ đụng tới).
  row=conn.execute(db.domains.select().filter_by(code=None,name=name)).mappings().fetchone()
  if not row: continue
  still_referenced=any(conn.execute(tbl.select().where(col_ref==row['id'])).fetchone() for tbl,col in ref_tables for col_ref in [tbl.c[col]])
  if still_referenced: continue
  conn.execute(db.domains.delete().where(db.domains.c.id==row['id'])); removed['domains']+=1
 return removed
def run_import():
 init_database()
 with db.get_connection() as conn:
  with conn.begin():
   prep(conn)
   results=[import_sheet(conn,x) for x in CONFIGS]
   removed=cleanup_known_import_garbage(conn)
 ct=db.fetch_table_counts(TABLES)
 REPORT.write_text(make_report(results,ct,removed),encoding='utf-8'); return results
def main():
 rs=run_import(); print(f'Core import finished. Report: {REPORT}')
 for r in rs: print(f"{r['sheet']}: read={r['rows']}, imported={r['ok']}, skipped={r['skip']}")
if __name__=='__main__': main()
